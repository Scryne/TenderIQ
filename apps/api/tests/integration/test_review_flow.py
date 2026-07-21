"""Sprint 3.2 inceleme akışı: onay/düzeltme/red + toplu onay + yorum/geçmiş + export.

Bulgular DB'ye doğrudan tohumlanır (pipeline sahteleri gerekmez): sınanan şey
inceleme API sözleşmesidir — durum makinesi, AuditLog düzenleme geçmişi, RLS
izolasyonu, RBAC ve export çıktısının (docx/xlsx) kaynak referansları.
"""

from __future__ import annotations

import io
import os
import uuid
import zipfile

import pytest
from fastapi.testclient import TestClient

import tenderiq_worker.db as worker_db
from tenderiq_core.findings import GroundingResolution, RequirementKind, RiskCategory, RiskSeverity
from tenderiq_core.models import (
    Document,
    DocumentKind,
    DocumentStatus,
    Requirement,
    RiskFlag,
)
from tenderiq_core.models import ParsedElement as ParsedElementRow
from tenderiq_core.parsing.types import ElementKind, ParseSource
from tenderiq_core.security.tokens import create_access_token

pytestmark = pytest.mark.integration

QUOTE = "Yüklenici tüm maddeleri karşılamak zorundadır."


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _register_and_login(client: TestClient, *, slug: str, email: str) -> tuple[str, str]:
    register = client.post(
        "/api/v1/auth/register",
        json={"org_name": slug, "org_slug": slug, "email": email, "password": "sifre-12345"},
    )
    assert register.status_code == 201, register.text
    login = client.post("/api/v1/auth/login", json={"email": email, "password": "sifre-12345"})
    assert login.status_code == 200
    return register.json()["tenant_id"], login.json()["access_token"]


class ReviewEnv:
    """Tohumlanmış inceleme ortamı: kiracı + ihale + 2 gereksinim + 1 risk."""

    def __init__(self, client: TestClient) -> None:
        suffix = uuid.uuid4().hex[:8]
        self.client = client
        self.tenant_id, self.token = _register_and_login(
            client, slug=f"org-rev-{suffix}", email=f"rev-{suffix}@org.com"
        )
        tender = client.post(
            "/api/v1/tenders", json={"title": "İnceleme İhalesi"}, headers=_auth(self.token)
        )
        assert tender.status_code == 201
        self.tender_id = tender.json()["id"]

        tenant_uuid = uuid.UUID(self.tenant_id)
        tender_uuid = uuid.UUID(self.tender_id)
        document_id = uuid.uuid4()
        element_id = uuid.uuid4()
        self.requirement_id = uuid.uuid4()
        self.requirement2_id = uuid.uuid4()
        self.risk_id = uuid.uuid4()

        # Worker'ın sync engine'i bu testin DATABASE_URL'ini görsün.
        worker_db._engine = None
        worker_db._factory = None
        with worker_db.tenant_session(tenant_uuid) as session:
            session.add(
                Document(
                    id=document_id,
                    tenant_id=tenant_uuid,
                    tender_id=tender_uuid,
                    filename="şartname.pdf",
                    content_type="application/pdf",
                    storage_key=f"{self.tenant_id}/{self.tender_id}/{document_id}/sartname.pdf",
                    kind=DocumentKind.ADMINISTRATIVE,
                    status=DocumentStatus.UPLOADED,
                    size_bytes=1234,
                )
            )
            session.add(
                ParsedElementRow(
                    id=element_id,
                    tenant_id=tenant_uuid,
                    document_id=document_id,
                    seq=0,
                    page=4,
                    kind=ElementKind.PARAGRAPH,
                    source=ParseSource.DIGITAL,
                    text=QUOTE,
                    section="Madde 7.2",
                    bbox_x0=72.0,
                    bbox_y0=120.0,
                    bbox_x1=480.0,
                    bbox_y1=150.0,
                )
            )
            common = {
                "tenant_id": tenant_uuid,
                "tender_id": tender_uuid,
                "document_id": document_id,
                "source_element_id": element_id,
                "grounding_resolution": GroundingResolution.ELEMENT,
                "source_quote": QUOTE,
            }
            session.add(
                Requirement(
                    id=self.requirement_id,
                    seq=0,
                    text="Yüklenici tüm maddeleri karşılamalıdır.",
                    kind=RequirementKind.ADMINISTRATIVE,
                    is_mandatory=True,
                    **common,
                )
            )
            session.add(
                Requirement(
                    id=self.requirement2_id,
                    seq=1,
                    text="Teminat mektubu sunulmalıdır.",
                    kind=RequirementKind.FINANCIAL,
                    is_mandatory=True,
                    **common,
                )
            )
            session.add(
                RiskFlag(
                    id=self.risk_id,
                    seq=0,
                    text="Gecikme cezası oranı yüksektir.",
                    severity=RiskSeverity.HIGH,
                    category=RiskCategory.PENALTY,
                    **common,
                )
            )


@pytest.fixture
def review_env(api_client: TestClient) -> ReviewEnv:
    return ReviewEnv(api_client)


def test_onay_duzeltme_gecmis_akisi(review_env: ReviewEnv) -> None:
    client, token = review_env.client, review_env.token
    rid = str(review_env.requirement_id)

    # Çıkarım sonrası her bulgu onay bekler; liste ucundan review alanı döner.
    listed = client.get(
        f"/api/v1/tenders/{review_env.tender_id}/requirements", headers=_auth(token)
    )
    assert listed.status_code == 200
    by_id = {row["id"]: row for row in listed.json()}
    assert by_id[rid]["review"] == {"status": "pending", "reviewed_by": None, "reviewed_at": None}

    # Onay: durum + kim/ne zaman dolar.
    approved = client.patch(
        f"/api/v1/requirements/{rid}", json={"action": "approve"}, headers=_auth(token)
    )
    assert approved.status_code == 200, approved.text
    review = approved.json()["review"]
    assert review["status"] == "approved"
    assert review["reviewed_by"] is not None
    assert review["reviewed_at"] is not None

    # Aynı eylem idempotenttir (yeni geçmiş kaydı üretmez).
    again = client.patch(
        f"/api/v1/requirements/{rid}", json={"action": "approve"}, headers=_auth(token)
    )
    assert again.status_code == 200
    assert again.json()["review"]["status"] == "approved"

    # action + içerik alanı birlikte verilemez; boş istek de geçersizdir.
    both = client.patch(
        f"/api/v1/requirements/{rid}",
        json={"action": "approve", "text": "x"},
        headers=_auth(token),
    )
    assert both.status_code == 400
    empty = client.patch(f"/api/v1/requirements/{rid}", json={}, headers=_auth(token))
    assert empty.status_code == 400

    # İçerik düzeltmesi: durum edited olur, yalnız değişen alan geçmişe yazılır.
    edited = client.patch(
        f"/api/v1/requirements/{rid}",
        json={"text": "Yüklenici tüm idari maddeleri karşılamalıdır.", "is_mandatory": True},
        headers=_auth(token),
    )
    assert edited.status_code == 200, edited.text
    assert edited.json()["review"]["status"] == "edited"
    assert edited.json()["text"] == "Yüklenici tüm idari maddeleri karşılamalıdır."

    # Düzenleme geçmişi (AuditLog): edit + approve kayıtları yeniden-eskiye.
    history = client.get(f"/api/v1/findings/requirement/{rid}/history", headers=_auth(token))
    assert history.status_code == 200
    entries = history.json()
    assert [entry["action"] for entry in entries] == ["finding.edited", "finding.approved"]
    changes = entries[0]["meta"]["changes"]
    assert set(changes) == {"text"}  # is_mandatory değişmedi → geçmişe girmez
    assert changes["text"]["to"] == "Yüklenici tüm idari maddeleri karşılamalıdır."

    # Geri alma: pending'e döner, kim/ne zaman temizlenir.
    reset = client.patch(
        f"/api/v1/requirements/{rid}", json={"action": "reset"}, headers=_auth(token)
    )
    assert reset.status_code == 200
    assert reset.json()["review"] == {
        "status": "pending",
        "reviewed_by": None,
        "reviewed_at": None,
    }

    # Risk PATCH'i de aynı sözleşmeyle çalışır (şiddet düzeltme örneği).
    risk_edit = client.patch(
        f"/api/v1/risks/{review_env.risk_id}",
        json={"severity": "medium"},
        headers=_auth(token),
    )
    assert risk_edit.status_code == 200
    assert risk_edit.json()["severity"] == "medium"
    assert risk_edit.json()["review"]["status"] == "edited"


def test_toplu_onay_red_ve_atlama(review_env: ReviewEnv) -> None:
    client, token = review_env.client, review_env.token
    rid1, rid2 = str(review_env.requirement_id), str(review_env.requirement2_id)

    # İlk gereksinim düzeltilmiş olsun: toplu onay EDITED'ı ezmemeli.
    edited = client.patch(
        f"/api/v1/requirements/{rid1}", json={"text": "Düzeltilmiş metin."}, headers=_auth(token)
    )
    assert edited.status_code == 200

    bogus = str(uuid.uuid4())
    bulk = client.post(
        f"/api/v1/tenders/{review_env.tender_id}/findings/bulk-review",
        json={
            "action": "approve",
            "items": [
                {"kind": "requirement", "id": rid1},
                {"kind": "requirement", "id": rid2},
                {"kind": "risk", "id": str(review_env.risk_id)},
                {"kind": "requirement", "id": bogus},
            ],
        },
        headers=_auth(token),
    )
    assert bulk.status_code == 200, bulk.text
    body = bulk.json()
    assert body == {"updated": 2, "unchanged": 1, "skipped": [bogus]}

    listed = client.get(
        f"/api/v1/tenders/{review_env.tender_id}/requirements", headers=_auth(token)
    ).json()
    statuses = {row["id"]: row["review"]["status"] for row in listed}
    assert statuses[rid1] == "edited"  # toplu onay düzeltildi izini korudu
    assert statuses[rid2] == "approved"

    # Toplu red açık iradedir: EDITED dahil hepsi reddedilir.
    reject = client.post(
        f"/api/v1/tenders/{review_env.tender_id}/findings/bulk-review",
        json={
            "action": "reject",
            "items": [
                {"kind": "requirement", "id": rid1},
                {"kind": "requirement", "id": rid2},
            ],
        },
        headers=_auth(token),
    )
    assert reject.status_code == 200
    assert reject.json()["updated"] == 2


def test_yorumlar_ve_izolasyon(review_env: ReviewEnv) -> None:
    client, token = review_env.client, review_env.token
    rid = str(review_env.requirement_id)

    created = client.post(
        f"/api/v1/findings/requirement/{rid}/comments",
        json={"body": "Bu maddeyi hukuk da görsün."},
        headers=_auth(token),
    )
    assert created.status_code == 201, created.text
    assert created.json()["body"] == "Bu maddeyi hukuk da görsün."
    assert created.json()["author_user_id"] is not None

    listed = client.get(f"/api/v1/findings/requirement/{rid}/comments", headers=_auth(token))
    assert listed.status_code == 200
    assert [comment["body"] for comment in listed.json()] == ["Bu maddeyi hukuk da görsün."]

    # Yorum, bulgunun geçmişinde de görünür.
    history = client.get(f"/api/v1/findings/requirement/{rid}/history", headers=_auth(token))
    assert "finding.commented" in [entry["action"] for entry in history.json()]

    # RLS: başka kiracı bulguyu göremez (PATCH/yorum → 404).
    suffix = uuid.uuid4().hex[:8]
    _other_tenant, other_token = _register_and_login(
        client, slug=f"org-diger-{suffix}", email=f"diger-{suffix}@org.com"
    )
    cross_patch = client.patch(
        f"/api/v1/requirements/{rid}", json={"action": "approve"}, headers=_auth(other_token)
    )
    assert cross_patch.status_code == 404
    cross_comments = client.get(
        f"/api/v1/findings/requirement/{rid}/comments", headers=_auth(other_token)
    )
    assert cross_comments.status_code == 404

    # RBAC: izleyici rolü inceleme yazamaz (403), okuyabilir (200).
    secret = os.environ["AUTH_SECRET"]
    viewer_token = create_access_token(
        user_id=uuid.uuid4(),
        tenant_id=uuid.UUID(review_env.tenant_id),
        role="viewer",
        secret=secret,
    )
    forbidden = client.patch(
        f"/api/v1/requirements/{rid}", json={"action": "approve"}, headers=_auth(viewer_token)
    )
    assert forbidden.status_code == 403
    viewer_read = client.get(
        f"/api/v1/findings/requirement/{rid}/history", headers=_auth(viewer_token)
    )
    assert viewer_read.status_code == 200


def test_export_word_excel_kaynak_referanslari(review_env: ReviewEnv) -> None:
    client, token = review_env.client, review_env.token
    rid = str(review_env.requirement_id)
    export_url = f"/api/v1/tenders/{review_env.tender_id}/export"

    # Onaylı bulgu yokken export 409 (insan-döngüde onay ön koşul).
    too_early = client.post(export_url, json={"format": "docx"}, headers=_auth(token))
    assert too_early.status_code == 409

    # include_pending ile onay beklemeden rapor alınabilir.
    pending_ok = client.post(
        export_url, json={"format": "docx", "include_pending": True}, headers=_auth(token)
    )
    assert pending_ok.status_code == 200

    # Bir gereksinim onaylanır, risk düzeltilir → rapor bu ikisini içerir.
    assert (
        client.patch(
            f"/api/v1/requirements/{rid}", json={"action": "approve"}, headers=_auth(token)
        ).status_code
        == 200
    )
    assert (
        client.patch(
            f"/api/v1/risks/{review_env.risk_id}",
            json={"text": "Gecikme cezası oranı sözleşme bedeline göre yüksektir."},
            headers=_auth(token),
        ).status_code
        == 200
    )

    docx = client.post(export_url, json={"format": "docx"}, headers=_auth(token))
    assert docx.status_code == 200, docx.text
    assert docx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml"
    )
    assert "attachment" in docx.headers["content-disposition"]
    assert docx.content[:2] == b"PK"  # OOXML = zip
    with zipfile.ZipFile(io.BytesIO(docx.content)) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8")
    # Kaynak referansı (doküman + sayfa no) ve onaylı içerik raporda görünür.
    assert "Kaynaklar" in document_xml
    assert "s. 4" in document_xml
    assert "Madde 7.2" in document_xml
    assert "Yüklenici tüm maddeleri karşılamalıdır." in document_xml
    assert "Gecikme cezası oranı sözleşme bedeline göre yüksektir." in document_xml
    # Onaylanmamış (pending) ikinci gereksinim rapora girmez.
    assert "Teminat mektubu sunulmalıdır." not in document_xml

    xlsx = client.post(export_url, json={"format": "xlsx"}, headers=_auth(token))
    assert xlsx.status_code == 200
    assert xlsx.content[:2] == b"PK"
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(xlsx.content))
    assert workbook.sheetnames[0] == "Özet"
    assert "Gereksinimler" in workbook.sheetnames
    requirement_rows = [
        [cell.value for cell in row] for row in workbook["Gereksinimler"].iter_rows()
    ]
    flattened = str(requirement_rows)
    assert "şartname.pdf" in flattened
    assert "Madde 7.2" in flattened  # kaynak kolonu her satırda
