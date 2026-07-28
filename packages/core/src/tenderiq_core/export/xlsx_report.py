"""Excel rapor üretimi (Sprint 3.2, §4.1) — openpyxl.

Şablon: "Özet" sayfası (firma + ihale + tarih + bölüm sayaçları) ve kategori
başına bir sayfa; her satırda içerik kolonları + İnceleme + kaynak kolonları
(Doküman, Sayfa, Bölüm, Alıntı) — sayfa no her satırda görünür (§4.1 kaynak
referansı). openpyxl yalnız bu modülde ve lazy import edilir.

**Formül enjeksiyonu savunması (CWE-1236).** Bu raporun hücreleri saldırgan
kontrolündeki metinden gelir: içerik, MÜŞTERİNİN YÜKLEDİĞİ üçüncü-taraf
şartnameden çıkarılır ve alıntılar birebir taşınır. openpyxl, ``=`` ile başlayan
bir string'i sessizce FORMÜL olarak işaretler (``data_type='f'``); böyle bir
hücre, raporu açan satın alma yetkilisinin makinesinde değerlendirilir
(``=cmd|'/c ...'!A0`` gibi DDE yükleri dâhil). Bu yüzden metin hücreleri açık
tiple yazılır (``_write_row``) — içerik DEĞİŞTİRİLMEZ, yalnız tipi sabitlenir.
Kesme işareti eklemek gösterimi bozardı ve bu ürün alıntı sadakati üzerine kurulu.
"""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from tenderiq_core.export.report import (
    REVIEW_STATUS_LABELS,
    TenderReport,
    truncate_quote,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# Excel sayfa adında yasak karakterler ve 31 karakter sınırı vardır.
_SHEET_TITLE_FORBIDDEN = str.maketrans(dict.fromkeys("[]:*?/\\", "-"))


def _sheet_title(title: str) -> str:
    return title.translate(_SHEET_TITLE_FORBIDDEN)[:31]


def _write_row(sheet: Worksheet, values: list[object]) -> None:
    """Satırı ekler ve metin hücrelerinin tipini string'e sabitler.

    openpyxl (3.1) tip çıkarımı bir string'i iki durumda metin OLMAKTAN çıkarır:
    ``=`` ile başlıyorsa formül (``f``), ``#REF!``/``#N/A`` gibi bir hata koduysa
    hata hücresi (``e``). İkisi de yükleyicinin kontrolündeki metinden tetiklenir;
    değer korunur, yalnız tip geri alınır.
    """
    sheet.append(values)
    for cell in sheet[sheet.max_row]:
        if isinstance(cell.value, str) and cell.data_type != "s":
            cell.data_type = "s"


def build_xlsx_report(report: TenderReport) -> bytes:
    """Rapordan Excel (.xlsx) baytları üretir (saf — DB/IO yok)."""
    from openpyxl import Workbook  # opsiyonel bağımlılık (export extra) — lazy
    from openpyxl.styles import Font

    bold = Font(bold=True)
    workbook = Workbook()

    summary: Worksheet = workbook.active
    summary.title = "Özet"
    summary.append(["İhale Analiz Raporu"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    # Firma adı ve ihale başlığı da kullanıcı girdisidir → açık tiple yazılır.
    _write_row(summary, ["Firma", report.organization])
    _write_row(summary, ["İhale", report.tender_title])
    summary.append(["Oluşturma", report.generated_at.strftime("%d.%m.%Y %H:%M")])
    summary.append([])
    summary.append(["Bölüm", "Bulgu sayısı"])
    summary["A7"].font = bold
    summary["B7"].font = bold
    for section in report.sections:
        _write_row(summary, [section.title, len(section.items)])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 40

    for section in report.sections:
        if not section.items:
            continue
        sheet = workbook.create_sheet(_sheet_title(section.title))
        headers = (*section.headers, "İnceleme", "Doküman", "Sayfa", "Bölüm", "Alıntı")
        sheet.append(list(headers))
        for cell in sheet[1]:
            cell.font = bold
        for item in section.items:
            _write_row(
                sheet,
                [
                    *item.cells,
                    REVIEW_STATUS_LABELS[item.review_status],
                    item.source.document,
                    item.source.page,
                    (item.source.section or "").strip(),
                    truncate_quote(item.source.quote),
                ],
            )
        sheet.freeze_panes = "A2"
        # İlk kolon (ana metin) geniş, kaynak kolonları orta genişlikte.
        widths = [60, *([16] * (len(section.headers) - 1)), 14, 28, 8, 24, 60]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
